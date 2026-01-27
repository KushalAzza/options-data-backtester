#!/usr/bin/env python3
"""
Convert backtest_results.json to Excel with color coding
"""

import json
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def load_backtest_results(json_file: str) -> dict:
    """Load backtest results from JSON file"""
    with open(json_file, 'r') as f:
        return json.load(f)


def create_excel_with_colors(results_data, output_file: str):
    """Create Excel file with color coding
    
    Handles two formats:
    1. Dictionary with 'results' key and summary data
    2. List of trade results (calculate summary from list)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Backtest Results"
    
    # Define colors
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    profit_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
    loss_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red
    neutral_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Light yellow
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Handle both formats: dict with 'results' key or direct list
    if isinstance(results_data, list):
        # Direct list format - calculate summary from list
        results = results_data
        actual_trades = [r for r in results if r.get('entry_reason') not in ['VIX_THRESHOLD_EXCEEDED', 'VIX_EMA_SIGNAL_BLOCKED', 'EMA_NEUTRAL']]
        
        unique_dates = set(r.get('date', '') for r in results if r.get('date'))
        total_trading_days = len(unique_dates)
        total_trades = len(actual_trades)
        total_pnl = sum(r.get('total_pnl', 0) for r in actual_trades)
        
        winning_trades = sum(1 for r in actual_trades if r.get('total_pnl', 0) > 0)
        losing_trades = sum(1 for r in actual_trades if r.get('total_pnl', 0) < 0)
        
        max_profit = max((r.get('total_pnl', 0) for r in actual_trades), default=0)
        max_loss = min((r.get('total_pnl', 0) for r in actual_trades), default=0)
        
        average_pnl = total_pnl / total_trades if total_trades > 0 else 0
        
        # Calculate drawdown
        cumulative_pnl = 0
        max_cumulative = 0
        max_drawdown = 0
        max_drawdown_days = 0
        drawdown_start = None
        
        for r in actual_trades:
            cumulative_pnl += r.get('total_pnl', 0)
            if cumulative_pnl > max_cumulative:
                max_cumulative = cumulative_pnl
                drawdown_start = None
            else:
                drawdown = max_cumulative - cumulative_pnl
                if drawdown > max_drawdown:
                    max_drawdown = drawdown
                if drawdown_start is None:
                    drawdown_start = r.get('date')
        
        summary_dict = {
            "total_trading_days": total_trading_days,
            "total_trades": total_trades,
            "total_reentries": 0,  # EOD script doesn't have re-entries
            "total_pnl": total_pnl,
            "total_orders": total_trades * 2,  # CE + PE
            "per_order_charges": 0,  # Will be filled from config if needed
            "total_charges": 0,
            "net_pnl": total_pnl,
            "winning_trades": winning_trades,
            "losing_trades": losing_trades,
            "average_pnl": average_pnl,
            "max_profit": max_profit,
            "max_loss": max_loss,
            "max_drawdown": max_drawdown,
            "max_drawdown_days": max_drawdown_days,
        }
    else:
        # Dictionary format with summary data
        summary_dict = results_data
        results = results_data.get("results", [])
    
    # Summary sheet
    summary_ws = wb.create_sheet("Summary", 0)
    summary_data = [
        ["Metric", "Value"],
        ["Total Trading Days", summary_dict.get("total_trading_days", 0)],
        ["Total Trades", summary_dict.get("total_trades", 0)],
        ["Total Re-entries", summary_dict.get("total_reentries", 0)],
        ["Total P&L", summary_dict.get("total_pnl", 0)],
        ["Total Orders", summary_dict.get("total_orders", 0)],
        ["Per Order Charges", summary_dict.get("per_order_charges", 0)],
        ["Total Charges", summary_dict.get("total_charges", 0)],
        ["Net P&L (after charges)", summary_dict.get("net_pnl", 0)],
        ["Winning Trades", summary_dict.get("winning_trades", 0)],
        ["Losing Trades", summary_dict.get("losing_trades", 0)],
        ["Average P&L", summary_dict.get("average_pnl", 0)],
        ["Max Profit", summary_dict.get("max_profit", 0)],
        ["Max Loss", summary_dict.get("max_loss", 0)],
        ["Max Drawdown", summary_dict.get("max_drawdown", 0)],
        ["Max Drawdown Days", summary_dict.get("max_drawdown_days", 0)],
    ]
    
    for row_idx, row_data in enumerate(summary_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = summary_ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:  # Header row
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left" if col_idx == 1 else "right", vertical="center")
            cell.border = border
    
    # Adjust column widths for summary
    summary_ws.column_dimensions['A'].width = 25
    summary_ws.column_dimensions['B'].width = 20
    
    # Results sheet
    if not results:
        print("No results to export")
        wb.save(output_file)
        return
    
    # Define headers
    headers = [
        "Date", "Trade #", "Entry Time", "Exit Time", "Entry Reason",
        "Fast EMA (Entry)", "Slow EMA (Entry)", "Fast EMA (Exit)", "Slow EMA (Exit)",
        "Expiry Date", "VIX (Entry)", "VIX (Exit)",
        "Nifty Entry", "Nifty Exit",
        "CE Strike", "CE Entry Price", "CE Entry Time", "CE Exit Price", "CE Exit Time", "CE Exit Reason", "CE Stopped", "CE P&L",
        "PE Strike", "PE Entry Price", "PE Entry Time", "PE Exit Price", "PE Exit Time", "PE Exit Reason", "PE Stopped", "PE P&L",
        "Total P&L", "Cumulative P&L"
    ]
    
    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    
    # Write data with cumulative P&L
    cumulative_pnl = 0.0
    for row_idx, result in enumerate(results, 2):
        total_pnl = result.get("total_pnl", 0) or 0
        cumulative_pnl += total_pnl
        
        row_data = [
            result.get("date", ""),
            result.get("trade_number", ""),
            result.get("entry_time", ""),
            result.get("exit_time", ""),
            result.get("entry_reason", ""),
            result.get("fast_ema_at_entry"),
            result.get("slow_ema_at_entry"),
            result.get("fast_ema_at_exit"),
            result.get("slow_ema_at_exit"),
            result.get("expiry_date", ""),
            result.get("vix_at_entry"),
            result.get("vix_at_exit"),
            result.get("nifty_entry_price"),
            result.get("nifty_exit_price"),
            result.get("ce_strike"),
            result.get("ce_entry_price"),
            result.get("ce_entry_time", ""),
            result.get("ce_exit_price"),
            result.get("ce_exit_time", ""),
            result.get("ce_exit_reason", ""),
            result.get("ce_stopped", False),
            result.get("ce_pnl", 0),
            result.get("pe_strike"),
            result.get("pe_entry_price"),
            result.get("pe_entry_time", ""),
            result.get("pe_exit_price"),
            result.get("pe_exit_time", ""),
            result.get("pe_exit_reason", ""),
            result.get("pe_stopped", False),
            result.get("pe_pnl", 0),
            result.get("total_pnl", 0),
            cumulative_pnl,  # Cumulative P&L
        ]
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal="center" if col_idx <= 4 or col_idx in [10, 17, 19, 24, 26] else "right", vertical="center")
            
            # Color code based on P&L
            total_pnl = result.get("total_pnl", 0)
            if total_pnl is not None:
                if total_pnl > 0:
                    cell.fill = profit_fill
                elif total_pnl < 0:
                    cell.fill = loss_fill
                else:
                    cell.fill = neutral_fill
            
            # Format numbers
            if isinstance(value, (int, float)) and value is not None:
                if col_idx in [6, 7, 8, 9, 11, 12, 13, 14, 16, 18, 22, 23, 25, 27, 30, 31]:  # Price/EMA/P&L columns
                    cell.number_format = '#,##0.00'
                elif col_idx in [15, 22]:  # Strike columns
                    cell.number_format = '#,##0'
                elif col_idx in [21, 28]:  # Stopped (boolean)
                    cell.value = "Yes" if value else "No"
    
    # Adjust column widths
    column_widths = {
        'A': 12,  # Date
        'B': 8,   # Trade #
        'C': 18,  # Entry Time
        'D': 18,  # Exit Time
        'E': 15,  # Entry Reason
        'F': 15,  # Fast EMA Entry
        'G': 15,  # Slow EMA Entry
        'H': 15,  # Fast EMA Exit
        'I': 15,  # Slow EMA Exit
        'J': 12,  # Expiry Date
        'K': 12,  # VIX Entry
        'L': 12,  # VIX Exit
        'M': 12,  # Nifty Entry
        'N': 12,  # Nifty Exit
        'O': 10,  # CE Strike
        'P': 12,  # CE Entry Price
        'Q': 18,  # CE Entry Time
        'R': 12,  # CE Exit Price
        'S': 18,  # CE Exit Time
        'T': 15,  # CE Exit Reason
        'U': 10,  # CE Stopped
        'V': 12,  # CE P&L
        'W': 10,  # PE Strike
        'X': 12,  # PE Entry Price
        'Y': 18,  # PE Entry Time
        'Z': 12,  # PE Exit Price
        'AA': 18, # PE Exit Time
        'AB': 15, # PE Exit Reason
        'AC': 10, # PE Stopped
        'AD': 12, # PE P&L
        'AE': 12, # Total P&L
        'AF': 15, # Cumulative P&L
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Save file
    wb.save(output_file)
    print(f"Excel file created: {output_file}")


def main():
    json_file = "backtest_results.json"
    output_file = "backtest_results.xlsx"
    
    if not os.path.exists(json_file):
        print(f"Error: {json_file} not found")
        return
    
    print(f"Loading {json_file}...")
    results_data = load_backtest_results(json_file)
    
    print(f"Creating Excel file: {output_file}...")
    create_excel_with_colors(results_data, output_file)
    
    print("Done!")


if __name__ == "__main__":
    main()
